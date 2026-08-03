"""Exportación detallada de matrices procesadas de OD."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Sequence
import math

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..series import (
    catalogo_series,
    extraer_serie,
    filtrar_catalogo,
)


# ============================================================
# Utilidades generales
# ============================================================

def obtener_ids_experimentales(
    matrices: dict,
    orden_ids: Sequence[Any] | None = None,
) -> list[Any]:
    """
    Obtiene los IDs numéricos disponibles en el diccionario de matrices.

    Por defecto se ordenan numéricamente:

        0, 1, 2, ..., 24

    Parameters
    ----------
    matrices:
        Diccionario producido por ``procesar_matrices``.

    orden_ids:
        Orden personalizado de los IDs. Por ejemplo:

        ``[9, 1, 2, 3, 4]``

        Los IDs indicados que no existan serán ignorados.

    Returns
    -------
    list
        Lista ordenada de IDs experimentales. No incluye el blanco.
    """

    disponibles = [
        clave
        for clave in matrices
        if isinstance(
            clave,
            (
                int,
                float,
                np.integer,
                np.floating,
            ),
        )
    ]

    if orden_ids is None:
        return sorted(
            disponibles,
            key=float,
        )

    disponibles_set = set(disponibles)

    ordenados = [
        clave
        for clave in orden_ids
        if clave in disponibles_set
    ]

    restantes = [
        clave
        for clave in disponibles
        if clave not in ordenados
    ]

    restantes = sorted(
        restantes,
        key=float,
    )

    return ordenados + restantes


def formatear_media_std(
    media: Any,
    std: Any,
    decimales: int = 4,
    separador: str = " ",
) -> str:
    """
    Convierte una media y desviación estándar al formato:

        0.1234 (0.0056)

    Los valores no finitos se muestran como ``NA``.
    """

    try:
        media_float = float(media)
    except (TypeError, ValueError):
        return "NA"

    try:
        std_float = float(std)
    except (TypeError, ValueError):
        std_float = math.nan

    if not np.isfinite(media_float):
        return "NA"

    media_texto = f"{media_float:.{decimales}f}"

    if not np.isfinite(std_float):
        return f"{media_texto}{separador}(NA)"

    return (
        f"{media_texto}{separador}"
        f"({std_float:.{decimales}f})"
    )


def _nombre_fila(indice: int) -> str:
    """Convierte 0 en A, 1 en B, 26 en AA, etc."""

    nombre = ""
    numero = indice + 1

    while numero:
        numero, residuo = divmod(numero - 1, 26)
        nombre = chr(65 + residuo) + nombre

    return nombre


def _obtener_media_std(
    matrices: dict,
    identificador: Any,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Obtiene las matrices de media y desviación para un ID o el blanco.
    """

    if identificador == "blanco":
        if "blanco_promedio" not in matrices:
            raise KeyError(
                "No existe la llave 'blanco_promedio' en matrices."
            )

        contenido = matrices["blanco_promedio"]

    else:
        if identificador not in matrices:
            raise KeyError(
                f"No existe el ID {identificador!r} en matrices."
            )

        contenido = matrices[identificador]

    if not isinstance(contenido, dict):
        raise TypeError(
            f"El contenido de {identificador!r} debe ser "
            "un diccionario con 'mean' y 'std'."
        )

    if "mean" not in contenido or "std" not in contenido:
        raise KeyError(
            f"El contenido de {identificador!r} no tiene "
            "las llaves 'mean' y 'std'."
        )

    media = np.asarray(
        contenido["mean"],
        dtype=float,
    )

    std = np.asarray(
        contenido["std"],
        dtype=float,
    )

    if media.shape != std.shape:
        raise ValueError(
            f"Las matrices mean y std de {identificador!r} "
            f"tienen formas diferentes: {media.shape} y {std.shape}."
        )

    return media, std


# ============================================================
# Archivo 1: matrices separadas por ID
# ============================================================

def exportar_matrices_por_id(
    matrices: dict,
    ruta_excel: str | Path,
    orden_ids: Sequence[Any] | None = None,
    nombres_filas: Sequence[str] | None = None,
    nombres_columnas: Sequence[str] | None = None,
    decimales: int = 4,
    incluir_blanco: bool = True,
) -> Path:
    """
    Exporta todas las matrices procesadas, separadas claramente por ID.

    El archivo contiene tres hojas:

    ``Media_STD``
        Matrices apiladas verticalmente con valores en formato
        ``promedio (desviación)``.

    ``Medias``
        Tabla numérica larga con todas las medias.

    ``Desviaciones``
        Tabla numérica larga con todas las desviaciones estándar.

    Parameters
    ----------
    matrices:
        Diccionario producido por ``procesar_matrices``.

    ruta_excel:
        Ruta del archivo Excel de salida.

    orden_ids:
        Orden personalizado de los IDs. Si no se proporciona,
        se usa orden numérico ascendente.

    nombres_filas:
        Etiquetas de las filas de la placa. Por defecto A, B, C, etc.

    nombres_columnas:
        Nombres de las columnas promediadas. Por defecto:
        Bloque 1, Bloque 2, etc.

    decimales:
        Número de decimales mostrado.

    incluir_blanco:
        Indica si el blanco se agrega al final.

    Returns
    -------
    pathlib.Path
        Ruta del archivo generado.
    """

    ruta_excel = Path(ruta_excel)
    ruta_excel.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    ids = obtener_ids_experimentales(
        matrices,
        orden_ids=orden_ids,
    )

    identificadores: list[Any] = list(ids)

    if incluir_blanco:
        identificadores.append("blanco")

    if not identificadores:
        raise ValueError(
            "No existen matrices para exportar."
        )

    primera_media, _ = _obtener_media_std(
        matrices,
        identificadores[0],
    )

    numero_filas, numero_columnas = primera_media.shape

    if nombres_filas is None:
        nombres_filas = [
            _nombre_fila(i)
            for i in range(numero_filas)
        ]

    if nombres_columnas is None:
        nombres_columnas = [
            f"Bloque {i + 1}"
            for i in range(numero_columnas)
        ]

    if len(nombres_filas) != numero_filas:
        raise ValueError(
            "La cantidad de nombres_filas no coincide "
            "con las filas de las matrices."
        )

    if len(nombres_columnas) != numero_columnas:
        raise ValueError(
            "La cantidad de nombres_columnas no coincide "
            "con las columnas de las matrices."
        )

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Media_STD"

    # Estilos
    relleno_titulo = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    relleno_encabezado = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )
    relleno_blanco = PatternFill(
        "solid",
        fgColor="F4CCCC",
    )

    fuente_titulo = Font(
        color="FFFFFF",
        bold=True,
        size=12,
    )
    fuente_encabezado = Font(
        bold=True,
    )

    borde_fino = Side(
        style="thin",
        color="B7B7B7",
    )
    borde = Border(
        left=borde_fino,
        right=borde_fino,
        top=borde_fino,
        bottom=borde_fino,
    )

    fila_excel = 1
    registros_media = []
    registros_std = []

    for identificador in identificadores:
        media, std = _obtener_media_std(
            matrices,
            identificador,
        )

        if media.shape != (
            numero_filas,
            numero_columnas,
        ):
            raise ValueError(
                f"La matriz {identificador!r} tiene forma "
                f"{media.shape}; se esperaba "
                f"{(numero_filas, numero_columnas)}."
            )

        titulo = (
            "BLANCO"
            if identificador == "blanco"
            else f"ID {identificador}"
        )

        # Título de la matriz
        hoja.merge_cells(
            start_row=fila_excel,
            start_column=1,
            end_row=fila_excel,
            end_column=numero_columnas + 1,
        )

        celda_titulo = hoja.cell(
            fila_excel,
            1,
            titulo,
        )
        celda_titulo.fill = (
            relleno_blanco
            if identificador == "blanco"
            else relleno_titulo
        )
        celda_titulo.font = (
            Font(
                bold=True,
                size=12,
            )
            if identificador == "blanco"
            else fuente_titulo
        )
        celda_titulo.alignment = Alignment(
            horizontal="center",
        )

        fila_excel += 1

        # Encabezados
        hoja.cell(
            fila_excel,
            1,
            "Fila",
        )

        for j, nombre_columna in enumerate(
            nombres_columnas,
            start=2,
        ):
            hoja.cell(
                fila_excel,
                j,
                nombre_columna,
            )

        for celda in hoja[fila_excel][
            :numero_columnas + 1
        ]:
            celda.fill = relleno_encabezado
            celda.font = fuente_encabezado
            celda.border = borde
            celda.alignment = Alignment(
                horizontal="center",
            )

        fila_excel += 1

        # Datos de la matriz
        for i in range(numero_filas):
            hoja.cell(
                fila_excel,
                1,
                nombres_filas[i],
            )
            hoja.cell(
                fila_excel,
                1,
            ).font = fuente_encabezado

            for j in range(numero_columnas):
                texto = formatear_media_std(
                    media[i, j],
                    std[i, j],
                    decimales=decimales,
                )

                celda = hoja.cell(
                    fila_excel,
                    j + 2,
                    texto,
                )
                celda.alignment = Alignment(
                    horizontal="center",
                )
                celda.border = borde

                registros_media.append(
                    {
                        "id": identificador,
                        "fila": nombres_filas[i],
                        "columna": nombres_columnas[j],
                        "fila_indice": i,
                        "columna_indice": j,
                        "media": media[i, j],
                    }
                )

                registros_std.append(
                    {
                        "id": identificador,
                        "fila": nombres_filas[i],
                        "columna": nombres_columnas[j],
                        "fila_indice": i,
                        "columna_indice": j,
                        "desviacion_estandar": std[i, j],
                    }
                )

            fila_excel += 1

        # Espacio entre matrices
        fila_excel += 2

    # Anchos
    hoja.column_dimensions["A"].width = 14

    for columna in range(
        2,
        numero_columnas + 2,
    ):
        hoja.column_dimensions[
            get_column_letter(columna)
        ].width = 22

    hoja.freeze_panes = "B3"

 # Crear hojas numéricas directamente con openpyxl
    hoja_medias = libro.create_sheet(
        "Medias"
    )
    hoja_std = libro.create_sheet(
        "Desviaciones"
    )

    encabezados_media = [
        "id",
        "fila",
        "columna",
        "fila_indice",
        "columna_indice",
        "media",
    ]

    encabezados_std = [
        "id",
        "fila",
        "columna",
        "fila_indice",
        "columna_indice",
        "desviacion_estandar",
    ]

    hoja_medias.append(
        encabezados_media
    )
    hoja_std.append(
        encabezados_std
    )

    for registro in registros_media:
        hoja_medias.append(
            [
                registro[columna]
                for columna in encabezados_media
            ]
        )

    for registro in registros_std:
        hoja_std.append(
            [
                registro[columna]
                for columna in encabezados_std
            ]
        )

    for hoja_numerica in (
        hoja_medias,
        hoja_std,
    ):
        for celda in hoja_numerica[1]:
            celda.fill = relleno_titulo
            celda.font = fuente_titulo
            celda.alignment = Alignment(
                horizontal="center",
            )

        hoja_numerica.freeze_panes = "A2"

        for columna in range(
            1,
            hoja_numerica.max_column + 1,
        ):
            hoja_numerica.column_dimensions[
                get_column_letter(columna)
            ].width = 20

    libro.save(ruta_excel)

    return ruta_excel



# ============================================================
# Archivo 2: series organizadas por grupo
# ============================================================

def construir_tablas_series_por_grupo(
    config: dict,
    matrices: dict,
    tiempos: Sequence[Any] | None = None,
    tipo: str = "todos",
    incluir_ids: Iterable[str] | None = None,
    excluir_ids: Iterable[str] | None = None,
    decimales: int = 4,
) -> dict[str, pd.DataFrame]:
    """
    Construye tablas donde cada renglón representa una serie
    experimental y cada columna representa un ID o tiempo.

    No incluye el blanco.

    Returns
    -------
    dict
        Diccionario con cuatro tablas:

        - ``Media_STD``
        - ``Medias``
        - ``Desviaciones``
        - ``Formato_largo``
    """

    submatriz = int(
        config["submatriz"]
    )

    if tiempos is None:
        tiempos = obtener_ids_experimentales(
            matrices
        )
    else:
        tiempos = [
            tiempo
            for tiempo in tiempos
            if tiempo in matrices
        ]

    series = filtrar_catalogo(
        config,
        incluir_ids=incluir_ids,
        excluir_ids=excluir_ids,
        tipo=tipo,
    )

    if not series:
        raise ValueError(
            "No existen series que cumplan los filtros indicados."
        )

    columnas_metadata = [
        "grupo",
        "aceite",
        "id_serie",
        "nombre",
        "tipo",
        "concentracion",
        "fila",
        "columnas_originales",
    ]

    filas_formateadas = []
    filas_medias = []
    filas_std = []
    filas_largas = []

    for serie in series:
        datos = extraer_serie(
            matrices=matrices,
            serie=serie,
            submatriz=submatriz,
            tiempos=tiempos,
            normalizacion="ninguno",
        )

        metadata = {
            "grupo": serie.grupo,
            "aceite": serie.aceite,
            "id_serie": serie.id,
            "nombre": serie.nombre,
            "tipo": serie.tipo,
            "concentracion": serie.concentracion,
            "fila": serie.fila,
            "columnas_originales": ",".join(
                str(columna)
                for columna in serie.columnas_originales
            ),
        }

        fila_formateada = dict(metadata)
        fila_media = dict(metadata)
        fila_std = dict(metadata)

        for indice, tiempo in enumerate(
            datos["tiempo"]
        ):
            # Evitar mostrar 0.0 en lugar de 0
            if float(tiempo).is_integer():
                tiempo_columna = int(tiempo)
            else:
                tiempo_columna = tiempo

            nombre_columna = (
                f"ID {tiempo_columna}"
            )

            media = float(
                datos["mean"][indice]
            )
            std = float(
                datos["std"][indice]
            )
            sem = float(
                datos["sem"][indice]
            )
            n = float(
                datos["n"][indice]
            )

            fila_formateada[
                nombre_columna
            ] = formatear_media_std(
                media,
                std,
                decimales=decimales,
            )

            fila_media[
                nombre_columna
            ] = media

            fila_std[
                nombre_columna
            ] = std

            filas_largas.append(
                {
                    **metadata,
                    "id_tiempo": tiempo_columna,
                    "media": media,
                    "desviacion_estandar": std,
                    "error_estandar": sem,
                    "n": n,
                    "media_std": formatear_media_std(
                        media,
                        std,
                        decimales=decimales,
                    ),
                }
            )

        filas_formateadas.append(
            fila_formateada
        )
        filas_medias.append(
            fila_media
        )
        filas_std.append(
            fila_std
        )

    df_formateada = pd.DataFrame(
        filas_formateadas
    )
    df_medias = pd.DataFrame(
        filas_medias
    )
    df_std = pd.DataFrame(
        filas_std
    )
    df_largo = pd.DataFrame(
        filas_largas
    )

    # Ordenar respetando grupos y orden original de configuración
    orden_series = {
        serie.id: posicion
        for posicion, serie in enumerate(
            catalogo_series(config)
        )
    }

        def ordenar_tabla(
        dataframe: pd.DataFrame,
    ) -> pd.DataFrame:

        dataframe = dataframe.copy()

        dataframe["_orden"] = dataframe[
            "id_serie"
        ].map(orden_series)

        dataframe.sort_values(
            by=[
                "grupo",
                "_orden",
            ],
            na_position="last",
            inplace=True,
        )

        dataframe.drop(
            columns="_orden",
            inplace=True,
        )

        columnas_tiempos = [
            columna
            for columna in dataframe.columns
            if str(columna).startswith("ID ")
        ]

        return dataframe[
            columnas_metadata
            + columnas_tiempos
        ]

    df_formateada = ordenar_tabla(
        df_formateada
    )

    df_medias = ordenar_tabla(
        df_medias
    )

    df_std = ordenar_tabla(
        df_std
    )

    df_largo["_orden"] = df_largo[
        "id_serie"
    ].map(orden_series)

    df_largo.sort_values(
        by=[
            "grupo",
            "_orden",
            "id_tiempo",
        ],
        na_position="last",
        inplace=True,
    )

    df_largo.drop(
        columns="_orden",
        inplace=True,
    )

    return {
        "Media_STD": df_formateada,
        "Medias": df_medias,
        "Desviaciones": df_std,
        "Formato_largo": df_largo,
    }


def exportar_series_por_grupo(
    config: dict,
    matrices: dict,
    ruta_excel: str | Path,
    tiempos: Sequence[Any] | None = None,
    tipo: str = "todos",
    incluir_ids: Iterable[str] | None = None,
    excluir_ids: Iterable[str] | None = None,
    decimales: int = 4,
) -> Path:
    """
    Exporta las series organizadas por grupo.

    Cada renglón corresponde a una serie y cada columna a un ID
    experimental. El blanco no se incluye.
    """

    ruta_excel = Path(
        ruta_excel
    )
    ruta_excel.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    tablas = construir_tablas_series_por_grupo(
        config=config,
        matrices=matrices,
        tiempos=tiempos,
        tipo=tipo,
        incluir_ids=incluir_ids,
        excluir_ids=excluir_ids,
        decimales=decimales,
    )

    with pd.ExcelWriter(
        ruta_excel,
        engine="openpyxl",
    ) as escritor:
        for nombre_hoja, tabla in tablas.items():
            tabla.to_excel(
                escritor,
                sheet_name=nombre_hoja[:31],
                index=False,
            )

    # Dar formato después de escribir
    from openpyxl import load_workbook

    libro = load_workbook(
        ruta_excel
    )

    relleno_encabezado = PatternFill(
        "solid",
        fgColor="1F4E78",
    )
    fuente_encabezado = Font(
        color="FFFFFF",
        bold=True,
    )
    relleno_grupo = PatternFill(
        "solid",
        fgColor="D9EAF7",
    )

    for hoja in libro.worksheets:
        hoja.freeze_panes = "I2"
        hoja.auto_filter.ref = hoja.dimensions

        # Encabezados
        for celda in hoja[1]:
            celda.fill = relleno_encabezado
            celda.font = fuente_encabezado
            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # Anchos
        for columna in range(
            1,
            hoja.max_column + 1,
        ):
            letra = get_column_letter(
                columna
            )

            if columna <= 8:
                hoja.column_dimensions[
                    letra
                ].width = 20
            else:
                hoja.column_dimensions[
                    letra
                ].width = 18

        # Separar visualmente los grupos
        if hoja.max_row > 1:
            columna_grupo = 1
            grupo_anterior = None

            for fila in range(
                2,
                hoja.max_row + 1,
            ):
                grupo_actual = hoja.cell(
                    fila,
                    columna_grupo,
                ).value

                if (
                    grupo_actual
                    != grupo_anterior
                ):
                    for columna in range(
                        1,
                        hoja.max_column + 1,
                    ):
                        hoja.cell(
                            fila,
                            columna,
                        ).fill = relleno_grupo

                grupo_anterior = grupo_actual

    libro.save(
        ruta_excel
    )

    return ruta_excel
